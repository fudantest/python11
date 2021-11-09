#!/usr/bin/python3
# -*- coding: utf-8 -*-
# @Time     :2021/11/1 21:03
# @Author   :day day up up
# @File     :do_excel.py
from openpyxl import load_workbook
from tools import project_path
from tools.read_config import ReadConfig
class DoExcel:
    def __init__(self,file_name):
        self.file_name = file_name
    def get_header(self):
        #获取第一行的标题行
        wb = load_workbook(self.file_name)
        mode = eval(ReadConfig.get_config(project_path.case_config_path, 'MODE', 'mode'))
        for key in mode:  # key取到mode字典中的key值，就是sheet_name
            sheet = wb[key]
        header = []#存储标题行
        for j in range(1,sheet.max_column + 1):
            header.append(sheet.cell(1,j).value)
        return header

    def get_data(self):#通过嵌套循环获取excel单元格的值
        #1.打开excel
        wb = load_workbook(self.file_name)
        #2.定位表单
        test_data = []  # 通过append追加数据到列表中
        mode = eval(ReadConfig.get_config(project_path.case_config_path,'MODE','mode'))
        for key in mode:#key取到mode字典中的key值，就是sheet_name
            sheet = wb[key]
            #3.读取表单中的值
            if mode[key] == "all":
                header = self.get_header()#拿到header的值
                for i in range(2,sheet.max_row + 1):#第2、3、4、5
                    sub_data = {}#通过赋值把excel取出的值放到字典中
                    for j in range(1,sheet.max_column + 1):#第1~8列
                        sub_data[header[j-1]] = sheet.cell(i,j).value
                    sub_data["sheet_name"] = key
                    test_data.append(sub_data)
            else:
                for case_id in mode[key]:
                    for i in range(case_id + 1, sheet.max_row + 1):
                        sub_data = {}  # 通过赋值把excel取出的值放到字典中
                        for j in range(1, sheet.max_column + 1):
                            sub_data[header[j - 1]] = sheet.cell(case_id + 1, j).value
                    sub_data["sheet_name"] = key
                    test_data.append(sub_data)
        return test_data
    def write_back(self,sheet_name,i,result,TestResult):#excel写回数据
        wb = load_workbook(self.file_name)
        sheet = wb[sheet_name]
        sheet.cell(i,7).value = result
        sheet.cell(i,8).value = TestResult
        wb.save(self.file_name)#保存结果
if __name__ == '__main__':
    test_data = DoExcel(project_path.test_data_path).get_data()
    header = DoExcel(project_path.test_data_path).get_header()
    print(test_data)
    print(len(test_data))
    print(header)
