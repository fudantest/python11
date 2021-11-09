#!/usr/bin/python3
# -*- coding: utf-8 -*-
# @Time     :2021/10/24 16:05
# @Author   :day day up up
# @File     :do_excel.py
from openpyxl import load_workbook
class DoExcel:
    def __init__(self,file_name,sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name
    def get_header(self):
        #获取第一行的标题行
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        header = []#存储标题行
        for j in range(1,sheet.max_column + 1):
            header.append(sheet.cell(1,j).value)
        return header

    def get_data(self):#通过嵌套循环获取excel单元格的值
        #1.打开excel
        wb = load_workbook(self.file_name)
        #2.定位表单
        sheet = wb[self.sheet_name]
        #3.读取表单中的值
        header = self.get_header()#拿到header的值
        test_data = []#通过append追加数据到列表中

        for i in range(2,sheet.max_row + 1):
            sub_data = {}#通过赋值把excel取出的值放到字典中
            for j in range(1,sheet.max_column + 1):
                sub_data[header[j-1]] = sheet.cell(i,j).value
            test_data.append(sub_data)
        return test_data
if __name__ == '__main__':
    test_data = DoExcel("../../test_data/test_data.xlsx", "test").get_data()
    header = DoExcel("../../test_data/test_data.xlsx", "test").get_header()
    print(test_data)
    print(header)