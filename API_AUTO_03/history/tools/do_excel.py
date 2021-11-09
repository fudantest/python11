#!/usr/bin/python3
# -*- coding: utf-8 -*-
# @Time     :2021/10/24 16:05
# @Author   :day day up up
# @File     :do_excel.py
from openpyxl import load_workbook
class DoExcel:
    def __init__(self,file_name,sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name
    def get_data(self):
        #1.打开excel
        wb = load_workbook(self.file_name)
        #2.定位表单
        sheet = wb[self.sheet_name]
        #3.读取表单中的值
        test_data = []#通过append追加数据到列表中
        for i in range(1,sheet.max_row + 1):
            sub_data = {}#通过赋值把excel取出的值放到字典中
            sub_data["method"] = sheet.cell(i,1).value
            sub_data["url"] = sheet.cell(i,2).value
            sub_data["data"] = sheet.cell(i,3).value
            sub_data["expected"] = sheet.cell(i,4).value
            test_data.append(sub_data)
        return test_data
if __name__ == '__main__':
    test_data = DoExcel("../../test_data/test_data.xlsx", "test").get_data()
    print(test_data)

